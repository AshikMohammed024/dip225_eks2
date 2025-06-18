from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

prices = [
    row[10] for row in ws.iter_rows(min_row=2, values_only=True)
    if isinstance(row[8], str) and 'LaserJet' in row[8] and isinstance(row[10], (int, float))
]

average = int(sum(prices) / len(prices)) if prices else 0
print(average)
