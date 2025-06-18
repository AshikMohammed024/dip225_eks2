from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

count = sum(
    1 for row in ws.iter_rows(min_row=2, values_only=True)
    if isinstance(row[3], str) and row[3].startswith('Ain') and isinstance(row[11], (int, float)) and row[11] < 40
)

print(count)
